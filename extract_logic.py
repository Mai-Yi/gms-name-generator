# -*- coding: utf-8 -*-
import openpyxl
import json

wb = openpyxl.load_workbook(r'c:\Users\ROG\Desktop\GMSR中文名起名器.xlsx', data_only=False)
log = []

# 名字生成 sheet：公式
ws_main = wb['名字生成']
for row in ws_main.iter_rows(max_row=21, max_col=18):
    for c in row:
        if c.value and isinstance(c.value, str) and c.value.startswith('='):
            log.append(('FORMULA', c.coordinate, c.value))

# 对应字符表
ws_map = wb['对应字符表']
map_rows = list(ws_map.iter_rows(min_row=1, max_row=2, max_col=10, values_only=True))
header = list(map_rows[0]) if map_rows else []
log.append(('MAP_HEADER', header))
mapping = {}
for row in ws_map.iter_rows(min_row=2, max_col=5, values_only=True):
    if row[0] is not None and str(row[0]).strip():
        key = str(row[0]).strip()
        val = str(row[1]).strip() if row[1] is not None else ''
        if val:
            mapping[key] = val
log.append(('MAP_COUNT', len(mapping)))

# 拼音查询表
ws_py = wb['拼音查询表']
py_first = list(ws_py.iter_rows(min_row=1, max_row=1, max_col=20, values_only=True))
py_header = list(py_first[0]) if py_first else []
log.append(('PY_HEADER', py_header))
py_rows = list(ws_py.iter_rows(min_row=2, max_row=15, max_col=20, values_only=True))
log.append(('PY_SAMPLE', [[str(x) if x is not None else '' for x in r] for r in py_rows]))

with open('d:/Name/char_mapping.json', 'w', encoding='utf-8') as f:
    json.dump(mapping, f, ensure_ascii=False)

with open('d:/Name/extract_log.txt', 'w', encoding='utf-8') as f:
    for item in log:
        f.write(str(item) + '\n')

print('OK', len(mapping))

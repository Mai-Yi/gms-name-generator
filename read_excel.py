# -*- coding: utf-8 -*-
import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'c:\Users\ROG\Desktop\GMSR中文名起名器.xlsx', data_only=True)
out = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append([str(c) if c is not None else '' for c in row])
    out.append({'name': sheet_name, 'rows': rows})
with open('d:/Name/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print('Done. Sheets:', [s['name'] for s in out])

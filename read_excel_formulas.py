# -*- coding: utf-8 -*-
import openpyxl
import json

# 读取公式（不用 data_only）
wb = openpyxl.load_workbook(r'c:\Users\ROG\Desktop\GMSR中文名起名器.xlsx', data_only=False)
out = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(max_row=min(30, ws.max_row), max_col=min(20, ws.max_column)):
        cells = []
        for c in row:
            v = c.value
            if hasattr(c, 'data_type') and c.data_type == 'f' and v:
                cells.append(('F', str(v)))
            else:
                cells.append(('V', str(v) if v is not None else ''))
        rows.append(cells)
    out.append({'name': sheet_name, 'rows': rows})
with open('d:/Name/excel_formulas.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print('Done')

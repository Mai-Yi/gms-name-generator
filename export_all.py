# -*- coding: utf-8 -*-
import openpyxl
import json

wb = openpyxl.load_workbook(r'c:\Users\ROG\Desktop\GMSR中文名起名器.xlsx', data_only=True)

# 对应字符表：所有列，找出汉字列和游戏字符列
ws_map = wb['对应字符表']
map_rows = list(ws_map.iter_rows(max_row=ws_map.max_row, max_col=8, values_only=True))
# 按前两列导出：可能是 汉字 与 游戏字，或 十六进制 与 字符
byte_to_char = {}
char_to_game = {}
for row in map_rows[1:]:
    a, b = (row[0], row[1]) if len(row) >= 2 else (None, None)
    if a is None:
        continue
    sa, sb = str(a).strip(), str(b).strip() if b else ''
    if len(sa) == 2 and all(c in '0123456789ABCDEFabcdef' for c in sa):
        byte_to_char[sa.upper()] = sb
    elif len(sa) == 1 and sa:
        char_to_game[sa] = sb


with open('d:/Name/byte_to_char.json', 'w', encoding='utf-8') as f:
    json.dump(byte_to_char, f, ensure_ascii=False)

# 拼音查询表：第一列拼音，其余列汉字
ws_py = wb['拼音查询表']
py_data = {}
for row in ws_py.iter_rows(min_row=2, max_row=ws_py.max_row, max_col=ws_py.max_column, values_only=True):
    if not row or row[0] is None:
        continue
    pinyin = str(row[0]).strip().rstrip('：').rstrip(':')
    chars = [str(c).strip() for c in row[1:] if c and str(c).strip()]
    if pinyin and pinyin not in py_data:
        py_data[pinyin] = []
    for c in chars:
        if c and c not in py_data[pinyin]:
            py_data[pinyin].append(c)

with open('d:/Name/pinyin_chars.json', 'w', encoding='utf-8') as f:
    json.dump(py_data, f, ensure_ascii=False)
